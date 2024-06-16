import openpyxl
from main import extract_game_info
import sys
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

"""
update_spreadsheet.py

This script takes a filepath of a Rocket League match scoreboard (full screenshot 
with no cropping) and updates a spreadsheet of the scoreboard values. The spreadsheet 
is saved.

Author: Kyle Stewart
Date: 2024-06-16

Usage:
    python update_spreadsheet.py <image_path>
"""

SPREADSHEET_PATH = "C:/Users/kylve/OneDrive/Documents/RL_Scrims_Summer_2024.xlsx"

def update_spreadsheet(filepath):
    """
    Updates the Excel spreadsheet by inserting a row of data read from the provided 
    image.

    Args:
        filepath (str): The path to the image.
    """

    # load workbook
    workbook = openpyxl.load_workbook(SPREADSHEET_PATH)

    # select Games worksheet
    sheet = workbook["Games"]

    # find the current table
    table = next(iter(sheet._tables.values()))

    if table is None:
        print("Table could not be found")

    # extract the new row to be added
    new_row = extract_game_info(filepath)

    next_row = table.ref.split(':')[1][2:]
    if (int(next_row) > 1):
        next_row = int(next_row) + 1
    next_row = int(next_row)

    for col_num, value in enumerate(new_row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)

    # Update the table's reference
    end_col_letter = get_column_letter(len(table.tableColumns))
    table.ref = f"{table.ref.split(':')[0]}:{end_col_letter}{next_row}"

    # Save the workbook
    workbook.save("C:/Users/kylve/OneDrive/Documents/RL_Scrims_Summer_2024.xlsx")

if len(sys.argv) < 1:
    print("Invalid number of args given. Make sure you are only providing the path to the image you want scanned.")
    sys.exit(1)

update_spreadsheet(sys.argv[1])